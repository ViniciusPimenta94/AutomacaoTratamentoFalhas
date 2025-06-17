'''BIBLIOTECA UTILIZADAS'''
from datetime import date
from resolucaofalhas import Automacao
import json
import os
import openpyxl
import pandas as pd
import re
import shutil

class Falhas():
    # Variáveis usadas em diversas funções
    def __init__(self):
        # Pasta contendo arquivos zips que deram falhas
        self.path = 'C:/Users/G053/Desktop/FALHAS'
        self.stack = {}
        self.twm = {}
        self.parser = {}
        self.regex = {}
        Falhas.sheets(self)
    # Database de falhas do Google Sheets
    def sheets(self):
        # Arquivo Excel retirado do Elastic Search (Kibana)
        csvDataframe = pd.read_csv('C:/Users/G053/Downloads/Pesquisa Fatura NotOk.csv')
        # Dados principais
        transction = (csvDataframe['fields.elasticInvoiceMessage.TransactionId']).to_list()
        fornecedor = (csvDataframe['fields.elasticInvoiceMessage.Supplier']).to_list()
        cliente = (csvDataframe['fields.elasticInvoiceMessage.Customer']).to_list()
        description = (csvDataframe['fields.elasticInvoiceMessage.Description']).to_list()

        for i, value in enumerate(transction):
            self.stack[value] = description[i]
            self.twm[value] = cliente[i]
            self.parser[value] = fornecedor[i]

        # Salvar excel formatado com as falhas
        csvDataframe.drop(['fields.elasticInvoiceMessage.Description'], axis=1, inplace=True)
        header = ['Data', 'Fornecedor', 'Cliente', 'Processo', 'FileName', 'TransctionId']
        csvDataframe.to_excel(f'{self.path}/Falhas.xlsx', index = False, header=header)
        Falhas.backup(self)
    # Backup das falhas
    def backup(self): 
        pathBackup = 'C:/Users/G053/Desktop/Backup Falhas'
        try:
            today = date.today()
            shutil.copytree(self.path, f"{pathBackup}/{today.strftime(r'%d-%m-%Y')}", dirs_exist_ok=False)
        except FileExistsError:
            print('Backup já realizado')
        Falhas.pattern(self)
    # JSON com as Regex e os tipos de Erros
    def pattern(self):
        pattern = f'C:/Users/G053/Desktop/pattern.json'
        with open(pattern, encoding='utf-8') as arquivo:
            self.regex = json.load(arquivo)
        Falhas.organizar(self)
    # Organizando as falhas por pastas
    def organizar(self):
        erros = {}
        wb = openpyxl.load_workbook(f'{self.path}/Falhas.xlsx')
        ws = wb['Sheet1']

        for pattern in self.regex['Regex']:
            for fatura in self.stack.keys():
                match = re.findall(f'{pattern}', self.stack[fatura])
                try:
                    if not match[0] is None:
                        erros[fatura] = self.regex['Erros'][pattern]
                except KeyError:
                    pass
                except IndexError:
                    pass

                try:
                    if not match[0] is None:
                        os.makedirs(self.regex['Regex'][pattern], exist_ok=True)
                        try:
                            shutil.move(f'{self.path}/{fatura}.zip', self.regex['Regex'][pattern])
                        except FileNotFoundError:
                            # Arquivo não encontrado
                            pass
                        except shutil.Error:
                            # Arquivo já existe na pasta
                            pass
                except IndexError:
                    pass
        count = 2
        while count <= ws.max_row:
            fatura = ws.cell(row=count, column=6).value
            try:
                ws.cell(row=count, column=8).value = erros[fatura]
                if erros[fatura] == 'Id fornecedor diferente':
                    ws.cell(row=count, column=7).value = 'OK'
                elif erros[fatura] == 'Mês ref não é data valida':
                    ws.cell(row=count, column=7).value = 'OK'
                elif erros[fatura] == 'Impostos':
                    ws.cell(row=count, column=7).value = 'OK'
                elif erros[fatura] == 'Nota fiscal com +20 caracteres':
                    ws.cell(row=count, column=7).value = 'OK'                    
            except KeyError:
                pass
            count+=1
        wb.save(f'{self.path}/Falhas.xlsx')
        Automacao()
Falhas()
